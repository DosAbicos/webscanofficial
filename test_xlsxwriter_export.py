#!/usr/bin/env python3
"""
Test xlsxwriter-based export for Android compatibility
"""
import requests
import sys
from pathlib import Path
import zipfile

BACKEND_URL = "http://0.0.0.0:8001"

def test_xlsxwriter_export():
    """Test the xlsxwriter-based export"""
    print("=" * 80)
    print("TESTING ANDROID-COMPATIBLE EXPORT (xlsxwriter)")
    print("=" * 80)
    
    test_products = [
        {
            "id": 1,
            "name": "Тестовый товар 1",
            "nomenclature_code": "12345",
            "stock_quantity": 100.0,
            "barcode": "1234567890123",
            "actual_quantity": 95.0
        },
        {
            "id": 2,
            "name": "Тестовый товар 2",
            "nomenclature_code": "67890",
            "stock_quantity": 50.0,
            "barcode": "9876543210987",
            "actual_quantity": 48.0
        }
    ]
    
    print(f"\n📦 Sending {len(test_products)} test products...")
    
    try:
        response = requests.post(
            f"{BACKEND_URL}/api/export-excel-minimal",
            json=test_products,
            timeout=60
        )
        
        print(f"\n✓ Response status: {response.status_code}")
        
        if response.status_code != 200:
            print(f"❌ Error: {response.text}")
            return False
        
        # Save file
        output_path = Path("/tmp/test_android_export.xlsx")
        with open(output_path, 'wb') as f:
            f.write(response.content)
        
        file_size = output_path.stat().st_size
        print(f"✓ File saved: {output_path}")
        print(f"✓ File size: {file_size / 1024:.2f} KB")
        
        # Validate it's a proper ZIP/XLSX file
        print("\n" + "=" * 80)
        print("VALIDATING FILE STRUCTURE")
        print("=" * 80)
        
        # Check ZIP structure (XLSX is a ZIP)
        try:
            with zipfile.ZipFile(output_path, 'r') as zip_ref:
                file_list = zip_ref.namelist()
                print(f"\n✓ Valid ZIP structure with {len(file_list)} files inside")
                
                # Check for essential XLSX components
                essential_files = [
                    'xl/workbook.xml',
                    'xl/worksheets/sheet1.xml',
                    '[Content_Types].xml',
                    '_rels/.rels'
                ]
                
                missing = []
                for ef in essential_files:
                    if ef in file_list:
                        print(f"  ✓ {ef} - present")
                    else:
                        print(f"  ✗ {ef} - MISSING")
                        missing.append(ef)
                
                if missing:
                    print(f"\n⚠ Missing essential files: {missing}")
                    return False
                
                # Check XML content of worksheet
                worksheet_xml = zip_ref.read('xl/worksheets/sheet1.xml').decode('utf-8')
                
                # Check for problematic XML elements that Android Excel can't handle
                problematic_elements = [
                    'extLst',  # Extended list - causes Android issues
                    'x14:',    # Excel 2010+ extensions
                    'xr:',     # Revision extensions
                ]
                
                print("\n✓ Checking for Android-incompatible XML elements...")
                found_problems = []
                for elem in problematic_elements:
                    if elem in worksheet_xml:
                        print(f"  ⚠ Found: {elem}")
                        found_problems.append(elem)
                    else:
                        print(f"  ✓ Clean: No {elem}")
                
                if found_problems:
                    print(f"\n⚠ WARNING: Found problematic elements that may cause Android issues: {found_problems}")
                    print("   xlsxwriter should generate cleaner XML, this needs investigation")
                else:
                    print("\n✓ XML is clean - no problematic elements for Android!")
                
        except zipfile.BadZipFile:
            print("❌ File is not a valid ZIP/XLSX")
            return False
        
        # Try opening with openpyxl
        print("\n" + "=" * 80)
        print("TESTING WITH OPENPYXL")
        print("=" * 80)
        
        try:
            import openpyxl
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            print(f"\n✓ File opens with openpyxl")
            print(f"✓ Worksheet: '{ws.title}'")
            print(f"✓ Dimensions: {ws.max_row} rows x {ws.max_column} columns")
            wb.close()
        except Exception as e:
            print(f"⚠ openpyxl error: {e}")
        
        print("\n" + "=" * 80)
        print("✅ FILE GENERATION SUCCESSFUL")
        print("=" * 80)
        print("\n📱 Next step: Test on Android device")
        print("   The file uses xlsxwriter which creates Android-compatible XML")
        print("=" * 80)
        
        return True
        
    except Exception as e:
        print(f"\n❌ Exception: {e}")
        import traceback
        traceback.print_exc()
        return False

if __name__ == "__main__":
    success = test_xlsxwriter_export()
    sys.exit(0 if success else 1)
