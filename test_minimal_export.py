#!/usr/bin/env python3
"""
Test script for the new /api/export-excel-minimal endpoint
"""
import requests
import json
import sys
from pathlib import Path

# Backend URL
BACKEND_URL = "http://0.0.0.0:8001"

def test_minimal_export():
    """Test the minimal export endpoint"""
    print("=" * 80)
    print("Testing /api/export-excel-minimal endpoint")
    print("=" * 80)
    
    # Create test data with a few products
    test_products = [
        {
            "id": 1,
            "name": "Тестовый продукт 1",
            "nomenclature_code": "123456",
            "stock_quantity": 100.0,
            "barcode": "1234567890123",
            "actual_quantity": 95.0
        },
        {
            "id": 2,
            "name": "Тестовый продукт 2",
            "nomenclature_code": "654321",
            "stock_quantity": 50.0,
            "barcode": "9876543210987",
            "actual_quantity": 48.0
        }
    ]
    
    print(f"\n📦 Sending {len(test_products)} test products to endpoint...")
    
    try:
        # Call the endpoint
        response = requests.post(
            f"{BACKEND_URL}/api/export-excel-minimal",
            json=test_products,
            timeout=60
        )
        
        print(f"Response status: {response.status_code}")
        print(f"Response headers: {dict(response.headers)}")
        
        if response.status_code == 200:
            # Save the file
            output_path = Path("/tmp/test_minimal_export.xlsx")
            with open(output_path, 'wb') as f:
                f.write(response.content)
            
            print(f"\n✓ File saved to: {output_path}")
            print(f"✓ File size: {output_path.stat().st_size / 1024:.2f} KB")
            
            # Try to open and validate with openpyxl
            try:
                import openpyxl
                wb = openpyxl.load_workbook(output_path)
                ws = wb.active
                
                print(f"\n✓ File successfully opened with openpyxl")
                print(f"✓ Worksheet name: {ws.title}")
                print(f"✓ Max row: {ws.max_row}")
                print(f"✓ Max column: {ws.max_column}")
                
                # Check if data exists in expected columns (9 and 10 for barcode and quantity)
                print(f"\n📊 Checking for test data...")
                barcodes_found = 0
                quantities_found = 0
                
                for row in range(1, min(ws.max_row + 1, 100)):  # Check first 100 rows
                    barcode = ws.cell(row=row, column=9).value
                    quantity = ws.cell(row=row, column=10).value
                    
                    if barcode and str(barcode).strip():
                        barcodes_found += 1
                        print(f"  Row {row}: Barcode = {barcode}, Quantity = {quantity}")
                    
                    if quantity is not None:
                        quantities_found += 1
                
                print(f"\n✓ Found {barcodes_found} barcodes")
                print(f"✓ Found {quantities_found} quantities")
                
                wb.close()
                
                print("\n" + "=" * 80)
                print("✅ TEST PASSED: File is valid and readable")
                print("=" * 80)
                return True
                
            except Exception as e:
                print(f"\n❌ Error opening file with openpyxl: {e}")
                print("=" * 80)
                print("❌ TEST FAILED: File is corrupted or invalid")
                print("=" * 80)
                return False
        else:
            print(f"\n❌ Error response: {response.text}")
            print("=" * 80)
            print("❌ TEST FAILED: Endpoint returned error")
            print("=" * 80)
            return False
            
    except Exception as e:
        print(f"\n❌ Error calling endpoint: {e}")
        import traceback
        traceback.print_exc()
        print("=" * 80)
        print("❌ TEST FAILED: Exception occurred")
        print("=" * 80)
        return False

if __name__ == "__main__":
    success = test_minimal_export()
    sys.exit(0 if success else 1)
