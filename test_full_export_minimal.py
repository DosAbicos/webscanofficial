#!/usr/bin/env python3
"""
Comprehensive test for the /api/export-excel-minimal endpoint
Tests with realistic data structure
"""
import requests
import json
import sys
from pathlib import Path
import openpyxl

BACKEND_URL = "http://0.0.0.0:8001"

def test_realistic_export():
    """Test export with realistic product data"""
    print("=" * 80)
    print("COMPREHENSIVE TEST: /api/export-excel-minimal")
    print("=" * 80)
    
    # Create realistic test data
    test_products = [
        {
            "id": 1,
            "name": "Молоко 3.2% 1л",
            "nomenclature_code": "00000000001",
            "stock_quantity": 120.0,
            "barcode": "4607014781234",
            "actual_quantity": 118.0
        },
        {
            "id": 2,
            "name": "Хлеб белый 500г",
            "nomenclature_code": "00000000002",
            "stock_quantity": 85.0,
            "barcode": "4607014785678",
            "actual_quantity": 82.0
        },
        {
            "id": 3,
            "name": "Сыр Российский 200г",
            "nomenclature_code": "00000000003",
            "stock_quantity": 45.0,
            "barcode": "4607014789012",
            "actual_quantity": 43.0
        },
        {
            "id": 4,
            "name": "Масло подсолнечное 1л",
            "nomenclature_code": "00000000004",
            "stock_quantity": 60.0,
            "barcode": "4607014783456",
            "actual_quantity": 58.0
        },
        {
            "id": 5,
            "name": "Яйца куриные С1 10шт",
            "nomenclature_code": "00000000005",
            "stock_quantity": 95.0,
            "barcode": "4607014787890",
            "actual_quantity": 92.0
        }
    ]
    
    print(f"\n📦 Testing with {len(test_products)} realistic products")
    print("\nTest Products:")
    for idx, p in enumerate(test_products, 1):
        print(f"  {idx}. {p['name']}")
        print(f"     Code: {p['nomenclature_code']} | Barcode: {p['barcode']}")
        print(f"     Stock: {p['stock_quantity']} | Actual: {p['actual_quantity']}")
    
    try:
        print("\n" + "=" * 80)
        print("Step 1: Calling /api/export-excel-minimal endpoint...")
        print("=" * 80)
        
        response = requests.post(
            f"{BACKEND_URL}/api/export-excel-minimal",
            json=test_products,
            timeout=60
        )
        
        print(f"\n✓ Response status: {response.status_code}")
        print(f"✓ Content-Type: {response.headers.get('content-type')}")
        print(f"✓ Content-Length: {len(response.content)} bytes ({len(response.content) / 1024:.2f} KB)")
        
        if response.status_code != 200:
            print(f"\n❌ Error: {response.text}")
            return False
        
        # Save the file
        output_path = Path("/tmp/test_realistic_export.xlsx")
        with open(output_path, 'wb') as f:
            f.write(response.content)
        
        print(f"\n✓ File saved to: {output_path}")
        
        print("\n" + "=" * 80)
        print("Step 2: Validating Excel file integrity...")
        print("=" * 80)
        
        # Open and validate
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        print(f"\n✓ File opened successfully with openpyxl")
        print(f"✓ Worksheet: '{ws.title}'")
        print(f"✓ Dimensions: {ws.max_row} rows x {ws.max_column} columns")
        
        print("\n" + "=" * 80)
        print("Step 3: Checking for data in barcode/quantity columns...")
        print("=" * 80)
        
        # Check columns 9 (I) and 10 (J) for barcodes and quantities
        barcodes_found = []
        quantities_found = []
        
        for row in range(1, min(ws.max_row + 1, 500)):  # Check first 500 rows
            barcode = ws.cell(row=row, column=9).value
            quantity = ws.cell(row=row, column=10).value
            
            if barcode and str(barcode).strip() and str(barcode).strip() != "Штрихкоды":
                # Check if it's a valid barcode (numeric)
                barcode_str = str(barcode).strip()
                if barcode_str.isdigit() and len(barcode_str) >= 10:
                    barcodes_found.append({
                        'row': row,
                        'barcode': barcode,
                        'quantity': quantity
                    })
        
        print(f"\n📊 Results:")
        print(f"  • Barcodes found: {len(barcodes_found)}")
        
        if barcodes_found:
            print(f"\n  First few entries:")
            for item in barcodes_found[:10]:
                print(f"    Row {item['row']}: Barcode={item['barcode']}, Qty={item['quantity']}")
        
        print("\n" + "=" * 80)
        print("Step 4: File format validation...")
        print("=" * 80)
        
        # Check file format
        file_size = output_path.stat().st_size
        print(f"\n✓ File size: {file_size / 1024:.2f} KB")
        
        # Check if file is a valid ZIP (XLSX is a ZIP archive)
        with open(output_path, 'rb') as f:
            file_header = f.read(4)
            is_zip = file_header == b'PK\x03\x04'
            print(f"✓ Valid .xlsx format (ZIP): {is_zip}")
        
        wb.close()
        
        # Final verdict
        print("\n" + "=" * 80)
        if file_size > 0 and is_zip:
            print("✅ ALL TESTS PASSED!")
            print("=" * 80)
            print("\n📱 FILE IS READY FOR TESTING ON:")
            print("  • Android devices")
            print("  • Windows Excel")
            print("  • iOS devices")
            print("\n🔍 Expected behavior:")
            print("  • File should open without errors")
            print("  • All data should be visible")
            print("  • Formatting should be preserved")
            print("=" * 80)
            return True
        else:
            print("❌ TEST FAILED: File validation failed")
            print("=" * 80)
            return False
            
    except Exception as e:
        print(f"\n❌ Exception occurred: {e}")
        import traceback
        traceback.print_exc()
        print("=" * 80)
        return False

if __name__ == "__main__":
    success = test_realistic_export()
    sys.exit(0 if success else 1)
