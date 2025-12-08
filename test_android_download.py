#!/usr/bin/env python3
"""
Test Android download - create minimal valid .xlsx file
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def create_minimal_xlsx():
    """Create a minimal but complete .xlsx file with test data"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventory"
    
    # Headers
    headers = ["Номенклатура", "Код", "Количество", "БУ", "Дебет", "Кредит", 
               "Дебет Остаток", "Кредит Остаток", "Штрихкоды", "Кол-во пофакту"]
    
    # Write headers (row 5)
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    
    # Test data
    test_products = [
        {
            "name": "IDEAL ПАРКЕТНЫЙ ЛАК ГЛЯНЦЕВЫЙ (10L)",
            "code": "20002755",
            "quantity": 15.0,
            "barcode": "TEST001",
            "actual_qty": 12.5
        },
        {
            "name": "IDEAL ПАРКЕТНЫЙ ЛАК ГЛЯНЦЕВЫЙ (2,5L)",
            "code": "20002659",
            "quantity": 25.0,
            "barcode": "TEST002",
            "actual_qty": 20.0
        }
    ]
    
    # Write data starting from row 10
    current_row = 10
    for product in test_products:
        # Product name row
        ws.cell(row=current_row, column=1, value=product["name"])
        ws.cell(row=current_row, column=3, value=product["quantity"])
        ws.cell(row=current_row, column=9, value=product["barcode"])
        ws.cell(row=current_row, column=10, value=product["actual_qty"])
        
        # "Кол." row
        ws.cell(row=current_row+1, column=2, value="Кол.")
        ws.cell(row=current_row+1, column=3, value=product["quantity"])
        
        # Code row
        ws.cell(row=current_row+2, column=1, value=product["code"])
        ws.cell(row=current_row+2, column=9, value=product["barcode"])
        ws.cell(row=current_row+2, column=10, value=product["actual_qty"])
        
        current_row += 4  # Space between products
    
    # Save
    output_path = '/tmp/minimal_test.xlsx'
    wb.save(output_path)
    
    print(f"✅ Created minimal test file: {output_path}")
    
    # Verify
    wb_check = openpyxl.load_workbook(output_path)
    ws_check = wb_check.active
    print(f"   Rows: {ws_check.max_row}, Columns: {ws_check.max_column}")
    
    # Check data
    for row in range(1, min(20, ws_check.max_row + 1)):
        col1 = ws_check.cell(row=row, column=1).value
        col9 = ws_check.cell(row=row, column=9).value
        col10 = ws_check.cell(row=row, column=10).value
        if col9:
            print(f"   Row {row}: {col1[:40] if col1 else ''} | {col9} | {col10}")
    
    print("\n✅ File is valid and ready for Android!")
    print("   Try opening this file on Android")
    print("   If this works but full export doesn't, the issue is file size/complexity")
    
    import os
    size_kb = os.path.getsize(output_path) / 1024
    print(f"\n   File size: {size_kb:.1f} KB")


if __name__ == '__main__':
    create_minimal_xlsx()
