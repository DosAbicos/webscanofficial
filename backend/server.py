from fastapi import FastAPI, APIRouter
from fastapi.responses import FileResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone
import xlrd
import xlwt
from xlutils.copy import copy as xl_copy


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")


# Define Models
class StatusCheck(BaseModel):
    model_config = ConfigDict(extra="ignore")  # Ignore MongoDB's _id field
    
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    client_name: str
    timestamp: datetime = Field(default_factory=lambda: datetime.now(timezone.utc))

class StatusCheckCreate(BaseModel):
    client_name: str

# Add your routes to the router instead of directly to app
@api_router.get("/")
async def root():
    return {"message": "Hello World"}

@api_router.post("/status", response_model=StatusCheck)
async def create_status_check(input: StatusCheckCreate):
    status_dict = input.model_dump()
    status_obj = StatusCheck(**status_dict)
    
    # Convert to dict and serialize datetime to ISO string for MongoDB
    doc = status_obj.model_dump()
    doc['timestamp'] = doc['timestamp'].isoformat()
    
    _ = await db.status_checks.insert_one(doc)
    return status_obj

@api_router.get("/status", response_model=List[StatusCheck])
async def get_status_checks():
    # Exclude MongoDB's _id field from the query results
    status_checks = await db.status_checks.find({}, {"_id": 0}).to_list(1000)
    
    # Convert ISO string timestamps back to datetime objects
    for check in status_checks:
        if isinstance(check['timestamp'], str):
            check['timestamp'] = datetime.fromisoformat(check['timestamp'])
    
    return status_checks


class Product(BaseModel):
    id: int
    name: str
    nomenclature_code: str
    stock_quantity: float
    barcode: Optional[str] = ""
    actual_quantity: Optional[float] = None


@api_router.post("/export-excel-debug")
async def export_excel_debug(products: List[Product]):
    """Debug endpoint - shows what would be exported without saving"""
    try:
        import xlrd
        
        original_path = '/app/sample_file.xls'
        rb = xlrd.open_workbook(original_path)
        original_sheet = rb.sheet_by_index(0)
        
        # Create product map by nomenclature_code (unique identifier)
        # This prevents issues with duplicate names
        product_map = {p.nomenclature_code: p for p in products if p.nomenclature_code}
        
        logger.info(f"Debug: Received {len(products)} products")
        
        # Find matches
        matches = []
        row_idx = 9
        
        while row_idx < original_sheet.nrows:
            cell_a = original_sheet.cell(row_idx, 0)
            cell_b = original_sheet.cell(row_idx + 1, 1) if row_idx + 1 < original_sheet.nrows else None
            
            if not cell_a.value:
                row_idx += 1
                continue
            
            cell_value = str(cell_a.value).strip()
            next_cell_value = str(cell_b.value).strip() if cell_b else ''
            
            if next_cell_value == 'Кол.':
                clean_name = cell_value.replace(' ', '')
                is_code = clean_name.isdigit()
                
                if not is_code and cell_value and cell_value != 'Итого':
                    # Get nomenclature code from 2 rows down
                    nomenclature_code = ''
                    if row_idx + 2 < original_sheet.nrows:
                        code_cell = original_sheet.cell(row_idx + 2, 0)
                        potential_code = str(code_cell.value).strip()
                        clean_code = potential_code.replace(' ', '')
                        if clean_code.isdigit():
                            nomenclature_code = potential_code
                    
                    if nomenclature_code and nomenclature_code in product_map:
                        product = product_map[nomenclature_code]
                        matches.append({
                            'row': row_idx + 1,
                            'excel_name': cell_value,
                            'excel_code': nomenclature_code,
                            'db_name': product.name,
                            'db_code': product.nomenclature_code,
                            'barcode': product.barcode,
                            'actual_qty': product.actual_quantity,
                            'match_type': 'by_code'
                        })
                
                row_idx += 2
            else:
                row_idx += 1
        
        # Also find products that didn't match
        matched_codes = {m['db_code'] for m in matches}
        unmatched = [p for p in products if p.nomenclature_code not in matched_codes]
        
        return {
            'total_received': len(products),
            'matches_found': len(matches),
            'unmatched': len(unmatched),
            'matches': matches[:20],  # First 20
            'unmatched_products': [{'name': p.name[:50], 'code': p.nomenclature_code, 'barcode': p.barcode} for p in unmatched[:10]]
        }
    except Exception as e:
        logger.error(f"Debug error: {e}")
        return {"error": str(e)}


@api_router.post("/export-excel-minimal")
async def export_excel_minimal(products: List[Product]):
    """
    Generate a clean .xlsx file from scratch using openpyxl.
    This endpoint creates a simpler file structure that should be more compatible
    with Android and Windows Excel readers.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
        
        logger.info(f"Starting minimal Excel export with {len(products)} products")
        
        # Read original file to extract structure and data
        original_path = '/app/frontend/public/sample_file.xls'
        rb = xlrd.open_workbook(original_path)
        original_sheet = rb.sheet_by_index(0)
        
        # Create new workbook from scratch
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Inventory"
        
        # Create product map by nomenclature_code
        product_map = {p.nomenclature_code: p for p in products if p.nomenclature_code}
        
        # Copy all data from original file first
        logger.info(f"Copying {original_sheet.nrows} rows from original file...")
        for row_idx in range(original_sheet.nrows):
            for col_idx in range(original_sheet.ncols):
                cell = original_sheet.cell(row_idx, col_idx)
                if cell.value is not None and cell.value != '':
                    ws.cell(row=row_idx+1, column=col_idx+1, value=cell.value)
        
        # Now update products data (barcodes and quantities)
        updated_count = 0
        row_idx = 9  # Start from row 10 (Excel uses 1-based indexing)
        
        while row_idx < original_sheet.nrows:
            try:
                cell_a = original_sheet.cell(row_idx, 0)
                cell_b = original_sheet.cell(row_idx + 1, 1) if row_idx + 1 < original_sheet.nrows else None
                
                if not cell_a.value:
                    row_idx += 1
                    continue
                
                cell_value = str(cell_a.value).strip()
                next_cell_value = str(cell_b.value).strip() if cell_b else ''
                
                if next_cell_value == 'Кол.':
                    clean_name = cell_value.replace(' ', '')
                    is_code = clean_name.isdigit()
                    
                    if not is_code and cell_value and cell_value != 'Итого':
                        # Get nomenclature code from 2 rows down
                        nomenclature_code = ''
                        if row_idx + 2 < original_sheet.nrows:
                            code_cell = original_sheet.cell(row_idx + 2, 0)
                            potential_code = str(code_cell.value).strip()
                            clean_code = potential_code.replace(' ', '')
                            if clean_code.isdigit():
                                nomenclature_code = potential_code
                        
                        # Match by nomenclature code
                        if nomenclature_code and nomenclature_code in product_map:
                            product = product_map[nomenclature_code]
                            
                            # Write barcode to column 9 (I in Excel, 0-indexed col 8)
                            if product.barcode:
                                ws.cell(row=row_idx+1, column=9, value=product.barcode)
                                # Also write to code row
                                if row_idx + 2 < original_sheet.nrows:
                                    ws.cell(row=row_idx+3, column=9, value=product.barcode)
                                logger.info(f"  ✓ Wrote barcode '{product.barcode}' for '{product.name[:40]}'")
                            
                            # Write actual quantity to column 10 (J in Excel, 0-indexed col 9)
                            if product.actual_quantity is not None:
                                ws.cell(row=row_idx+1, column=10, value=product.actual_quantity)
                                # Also write to code row
                                if row_idx + 2 < original_sheet.nrows:
                                    ws.cell(row=row_idx+3, column=10, value=product.actual_quantity)
                                logger.info(f"  ✓ Wrote quantity {product.actual_quantity} for '{product.name[:40]}'")
                            
                            updated_count += 1
                    
                    row_idx += 2
                else:
                    row_idx += 1
            except Exception as e:
                logger.error(f"Error processing row {row_idx}: {e}")
                row_idx += 1
        
        logger.info(f"Updated {updated_count} products in Excel")
        
        # Save to file with simple format
        timestamp = datetime.now().strftime("%Y-%m-%d")
        output_path = f'/tmp/inventory_export_{timestamp}.xlsx'
        wb.save(output_path)
        
        # Verify file was created
        import os
        if os.path.exists(output_path):
            file_size = os.path.getsize(output_path)
            logger.info(f"✓ Excel file created: {output_path} ({file_size / 1024:.1f} KB)")
        else:
            raise Exception("File was not created")
        
        return FileResponse(
            output_path,
            media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            filename=f'inventory_export_{timestamp}.xlsx',
            headers={
                'Content-Disposition': f'attachment; filename="inventory_export_{timestamp}.xlsx"',
                'Cache-Control': 'no-cache'
            }
        )
    
    except Exception as e:
        logger.error(f"Error in export_excel_minimal: {e}")
        import traceback
        traceback.print_exc()
        return {"error": str(e)}


@api_router.post("/export-excel")
async def export_excel(products: List[Product]):
    """
    Export Excel with all products, maintaining original formatting
    Data is written to the same row as product name (column 8 and 9)
    """
    try:
        # Load original file with formatting
        original_path = '/app/sample_file.xls'
        rb = xlrd.open_workbook(original_path, formatting_info=True)
        wb = xl_copy(rb)
        sheet = wb.get_sheet(0)
        
        # Create product map by nomenclature_code (unique identifier)
        # This prevents issues with duplicate names
        product_map = {p.nomenclature_code: p for p in products if p.nomenclature_code}
        
        # Update products in Excel
        row_idx = 9  # Start from row 10
        original_sheet = rb.sheet_by_index(0)
        
        logger.info(f"=" * 80)
        logger.info(f"Starting Excel export with {len(products)} products")
        logger.info(f"Product map has {len(product_map)} entries")
        
        # Log ALL product names for debugging
        logger.info(f"Products to export:")
        for i, p in enumerate(products):
            logger.info(f"  [{i+1}] '{p.name[:60]}' | Barcode: {p.barcode} | Qty: {p.actual_quantity}")
        
        # STEP 1: Clear all existing data in columns 8 and 9
        # This ensures deleted products are removed from export
        logger.info("Clearing old barcode and quantity data...")
        clear_row = 9
        cleared_count = 0
        
        while clear_row < original_sheet.nrows:
            try:
                # Clear barcode (column 8) and quantity (column 9)
                sheet.write(clear_row, 8, '')
                sheet.write(clear_row, 9, '')
                cleared_count += 1
                clear_row += 1
            except:
                break
        
        logger.info(f"Cleared {cleared_count} rows")
        
        # STEP 2: Write new data only for products with barcodes
        updated_count = 0
        row_idx = 9
        
        while row_idx < original_sheet.nrows:
            try:
                cell_a = original_sheet.cell(row_idx, 0)
                cell_b = original_sheet.cell(row_idx + 1, 1) if row_idx + 1 < original_sheet.nrows else None
                
                if not cell_a.value:
                    row_idx += 1
                    continue
                
                cell_value = str(cell_a.value).strip()
                next_cell_value = str(cell_b.value).strip() if cell_b else ''
                
                if next_cell_value == 'Кол.':
                    clean_name = cell_value.replace(' ', '')
                    is_code = clean_name.isdigit()
                    
                    if not is_code and cell_value and cell_value != 'Итого':
                        product = None
                        
                        # Get nomenclature code from 2 rows down
                        nomenclature_code = ''
                        if row_idx + 2 < original_sheet.nrows:
                            code_cell = original_sheet.cell(row_idx + 2, 0)
                            potential_code = str(code_cell.value).strip()
                            clean_code = potential_code.replace(' ', '')
                            if clean_code.isdigit():
                                nomenclature_code = potential_code
                        
                        # Match by nomenclature code (unique identifier)
                        if nomenclature_code and nomenclature_code in product_map:
                            product = product_map[nomenclature_code]
                            logger.info(f"  ✓ Matched by code: '{cell_value[:40]}' | Code: {nomenclature_code}")
                        else:
                            logger.debug(f"  ✗ No match for code: {nomenclature_code} | Name: '{cell_value[:40]}'")
                        
                        if not product:
                            row_idx += 2
                            continue
                        
                        # Write barcode to column 8 (same row as product name)
                        if product.barcode:
                            sheet.write(row_idx, 8, product.barcode)
                            # Also write to code row (2 rows down)
                            if row_idx + 2 < original_sheet.nrows:
                                sheet.write(row_idx + 2, 8, product.barcode)
                            logger.info(f"  ✓ Wrote barcode '{product.barcode}' at row {row_idx + 1}")
                        
                        # Write actual quantity to column 9 (same row as product name)
                        if product.actual_quantity is not None:
                            sheet.write(row_idx, 9, product.actual_quantity)
                            # Also write to code row (2 rows down)
                            if row_idx + 2 < original_sheet.nrows:
                                sheet.write(row_idx + 2, 9, product.actual_quantity)
                            logger.info(f"  ✓ Wrote quantity {product.actual_quantity} at row {row_idx + 1}")
                        
                        updated_count += 1
                    
                    row_idx += 2
                else:
                    row_idx += 1
            except Exception as e:
                logger.error(f"Error processing row {row_idx}: {e}")
                row_idx += 1
        
        logger.info(f"Updated {updated_count} products in Excel")
        
        # Save to temporary file (keep as .xls for now, but can convert to .xlsx if needed)
        timestamp = datetime.now().strftime("%Y-%m-%d")
        output_path = f'/tmp/updated_inventory_{timestamp}.xls'
        wb.save(output_path)
        
        logger.info(f"Excel file saved to {output_path}")
        
        # Convert to .xlsx for better compatibility with Android
        try:
            import openpyxl
            
            # Read .xls and convert to .xlsx
            old_wb = xlrd.open_workbook(output_path)
            old_sheet = old_wb.sheet_by_index(0)
            
            # Create new .xlsx workbook
            new_wb = openpyxl.Workbook()
            new_ws = new_wb.active
            
            # Copy all data - optimize by skipping empty cells
            logger.info(f"Converting {old_sheet.nrows} rows to .xlsx...")
            
            for row_idx in range(old_sheet.nrows):
                has_data = False
                for col_idx in range(old_sheet.ncols):
                    cell = old_sheet.cell(row_idx, col_idx)
                    if cell.value is not None and cell.value != '':
                        new_ws.cell(row=row_idx+1, column=col_idx+1, value=cell.value)
                        has_data = True
                
                # Progress log every 1000 rows
                if row_idx % 1000 == 0 and row_idx > 0:
                    logger.info(f"  Processed {row_idx} rows...")
            
            # Save as .xlsx
            xlsx_path = f'/tmp/updated_inventory_{timestamp}.xlsx'
            new_wb.save(xlsx_path)
            
            # Get file size
            import os
            file_size = os.path.getsize(xlsx_path)
            logger.info(f"Converted to .xlsx: {xlsx_path} ({file_size / 1024:.1f} KB)")
            
            return FileResponse(
                xlsx_path,
                media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                filename=f'updated_inventory_{timestamp}.xlsx'
            )
        except Exception as convert_error:
            logger.warning(f"Could not convert to .xlsx: {convert_error}, returning .xls")
            # Fallback to .xls if conversion fails
            return FileResponse(
                output_path,
                media_type='application/vnd.ms-excel',
                filename=f'updated_inventory_{timestamp}.xls'
            )
    
    except Exception as e:
        logger.error(f"Error in export_excel: {e}")
        import traceback
        traceback.print_exc()
        return {"error": str(e)}

# Include the router in the main app
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()